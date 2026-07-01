import streamlit as st
import pandas as pd
from datetime import date
import openpyxl
import io

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Trailer Status Dashboard",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    /* ── Global background ── */
    .stApp { background: #F8FAFC; }
    .main .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: #FFFFFF !important;
        border-right: 1px solid #E2E8F0 !important;
    }
    [data-testid="stSidebar"] h2 {
        color: #0F172A !important;
        font-size: 0.95rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.06em;
        text-transform: uppercase;
    }

    /* ── Header ── */
    .dash-header {
        background: linear-gradient(135deg, #1B4FD8 0%, #1e40af 50%, #1e3a8a 100%);
        border-radius: 16px;
        padding: 28px 36px;
        margin-bottom: 24px;
        position: relative;
        overflow: hidden;
        box-shadow: 0 8px 32px rgba(27,79,216,0.25);
    }
    .dash-header::before {
        content: '';
        position: absolute;
        top: -60px; right: -40px;
        width: 200px; height: 200px;
        background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
        border-radius: 50%;
    }
    .dash-header::after {
        content: '';
        position: absolute;
        bottom: -50px; left: 35%;
        width: 180px; height: 180px;
        background: radial-gradient(circle, rgba(255,255,255,0.07) 0%, transparent 70%);
        border-radius: 50%;
    }
    .dash-header-eyebrow {
        font-size: 0.68rem; font-weight: 600; letter-spacing: 0.18em;
        text-transform: uppercase; color: rgba(255,255,255,0.65);
        margin-bottom: 6px;
    }
    .dash-header h1 {
        margin: 0 0 6px; font-size: 1.9rem; font-weight: 800;
        color: #FFFFFF; letter-spacing: -0.02em; line-height: 1.15;
    }
    .dash-header p { margin: 0; color: rgba(255,255,255,0.7); font-size: 0.85rem; }
    .dash-header-badge {
        display: inline-flex; align-items: center; gap: 6px;
        background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25);
        border-radius: 20px; padding: 4px 12px;
        font-size: 0.72rem; font-weight: 600; color: #FFFFFF;
        letter-spacing: 0.05em; margin-top: 12px;
    }
    .dot { width: 6px; height: 6px; background: #34D399;
           border-radius: 50%; display: inline-block;
           box-shadow: 0 0 6px #34D399; }

    /* ── Metric cards ── */
    .metric-card {
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 14px;
        padding: 16px 18px;
        position: relative; overflow: hidden;
        box-shadow: 0 1px 6px rgba(0,0,0,0.04);
        transition: box-shadow 0.2s, transform 0.2s;
    }
    .metric-card:hover {
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        transform: translateY(-1px);
    }
    .metric-card .accent-bar {
        position: absolute; top: 0; left: 0; right: 0;
        height: 3px; border-radius: 14px 14px 0 0;
    }
    .metric-card .val {
        font-size: 2rem; font-weight: 800;
        letter-spacing: -0.03em; line-height: 1;
        margin-bottom: 4px;
    }
    .metric-card .lbl {
        font-size: 0.68rem; font-weight: 600;
        letter-spacing: 0.07em; text-transform: uppercase;
        color: #94A3B8;
    }

    /* ── Upload box ── */
    .upload-box {
        background: #FFFFFF;
        border: 1.5px solid #E2E8F0;
        border-radius: 12px;
        padding: 14px 16px 10px 16px;
        margin-bottom: 20px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }
    .upload-box h4 {
        margin: 0 0 2px;
        color: #0F172A !important;
        font-size: 0.95rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.01em;
    }
    .upload-box .upload-desc {
        font-size: 0.75rem;
        color: #64748B;
        margin-bottom: 8px;
        font-weight: 400;
    }
    /* ลด gap ระหว่าง desc กับ uploader */
    .upload-box [data-testid="stFileUploader"] {
        margin-top: -4px !important;
    }
    .upload-box [data-testid="stFileUploaderDropzone"] {
        padding: 10px !important;
        border-radius: 8px !important;
        border-color: #CBD5E1 !important;
        background: #F8FAFC !important;
    }

    /* ── Dataframe ── */
    [data-testid="stDataFrame"] {
        border-radius: 12px !important;
        overflow: hidden;
        border: 1px solid #E2E8F0 !important;
        box-shadow: 0 1px 6px rgba(0,0,0,0.04) !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        background: linear-gradient(135deg, #1B4FD8, #1e40af) !important;
        color: white !important; border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important; font-size: 0.85rem !important;
        letter-spacing: 0.03em !important;
        box-shadow: 0 4px 14px rgba(27,79,216,0.3) !important;
        transition: all 0.2s !important;
    }
    .stButton > button:hover {
        box-shadow: 0 6px 22px rgba(27,79,216,0.45) !important;
        transform: translateY(-1px) !important;
    }

    /* ── Count label ── */
    .count-label { font-size: 0.82rem; color: #475569; margin-bottom: 8px; font-weight: 500; }
    .count-label b { color: #1B4FD8; font-weight: 700; }

    /* ── Divider ── */
    hr { border-color: #E2E8F0 !important; }

    /* ── Expander ── */
    [data-testid="stExpander"] {
        background: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 10px !important;
    }
    [data-testid="stExpander"] summary {
        color: #0F172A !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
    }
    [data-testid="stExpander"] summary:hover {
        color: #1B4FD8 !important;
    }
    [data-testid="stExpander"] p,
    [data-testid="stExpander"] td,
    [data-testid="stExpander"] th {
        color: #1E293B !important;
        font-size: 0.85rem !important;
    }
    [data-testid="stExpander"] th {
        background-color: #F1F5F9 !important;
        font-weight: 700 !important;
    }

    /* ── Download button ── */
    .stDownloadButton > button {
        background: #EFF6FF !important;
        border: 1px solid #BFDBFE !important;
        color: #1B4FD8 !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
    }

    /* ── Scrollbar ── */
    ::-webkit-scrollbar { width: 4px; height: 4px; }
    ::-webkit-scrollbar-track { background: #F1F5F9; }
    ::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 4px; }

    /* ── Upload gap fix ── */
    .upload-box [data-testid="stFileUploader"] { margin-top: -8px !important; }
    .upload-box [data-testid="stFileUploaderDropzone"] {
        padding: 10px !important;
        border-radius: 8px !important;
        border-color: #E2E8F0 !important;
    }

    /* ── Warning fix ── */
    .custom-warning {
        background: #FFFBEB;
        border: 1px solid #F59E0B;
        border-left: 4px solid #F59E0B;
        border-radius: 8px;
        padding: 10px 14px;
        color: #92400E !important;
        font-size: 0.82rem;
        font-weight: 500;
        margin: 6px 0 10px;
    }
    .custom-success {
        background: #F0FDF4;
        border: 1px solid #22C55E;
        border-left: 4px solid #22C55E;
        border-radius: 8px;
        padding: 10px 14px;
        color: #166534 !important;
        font-size: 0.82rem;
        font-weight: 500;
        margin: 6px 0 10px;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def normalize(val):
    if pd.isna(val): return ""
    return str(val).strip().upper()

def find_header_row(ws, keyword, max_rows=15):
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        for cell in row:
            if cell and keyword.upper() in str(cell).upper(): return i
    return 0

def sheet_df(wb, sheet_name, keyword="Fleet"):
    ws = wb[sheet_name]
    hdr = find_header_row(ws, keyword)
    data = list(ws.values)
    if not data or hdr >= len(data): return pd.DataFrame()
    cols = [str(c).strip() if c is not None else "" for c in data[hdr]]
    return pd.DataFrame(data[hdr + 1:], columns=cols)

def find_col(df, candidates):
    for c in candidates:
        for col in df.columns:
            if c.upper() in col.upper(): return col
    return None

def is_date_pattern_sheet(name):
    months = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    lower = name.lower()
    return any(m in lower for m in months) and any(c.isdigit() for c in lower)

STATION_MAP = {
    "Sam Khok DC":                         "SKDC",
    "Wang Noi Ambient DC":                 "WNDC",
    "Surat Thani Regional DC":             "SRRDC",
    "ศูนย์กระจายสินค้าโลตัสลำพูน":        "LPRDC",
    "Lumlukka DC":                         "LLKDC",
    "Bang Bua Tong DC":                    "BBTDC",
    "Khon-Kaen Regional Ambient-Fresh DC": "KKRDC",
    "Area CCS2":                           "Uni",
}

# ─────────────────────────────────────────────────────────────────────────────
# CORE LOGIC — ตรงกับ app_dc_test.py 100%
# ─────────────────────────────────────────────────────────────────────────────

def build_lookups(wb_vor_all, wb_vor_lin, wb_trucker, today):
    """อ่านไฟล์ทั้งหมด 1 ครั้ง สร้าง set/dict สำหรับค้นหา O(1)"""
    # VOR set
    vor_set = set()
    if "VOR Report" in wb_vor_all.sheetnames:
        df = sheet_df(wb_vor_all, "VOR Report", "Fleet")
        col = find_col(df, ["Fleet No", "FleetNo", "Fleet#"])
        if col: vor_set |= set(df[col].map(normalize).values) - {""}

    # Dispose set
    dispose_set = set()
    for sname in wb_vor_all.sheetnames:
        if "รอขาย" not in sname: continue
        df = sheet_df(wb_vor_all, sname, "Fleet")
        col = find_col(df, ["Fleet No", "FleetNo", "Fleet#"])
        if col: dispose_set |= set(df[col].map(normalize).values) - {""}

    for sname in [s for s in wb_vor_lin.sheetnames if is_date_pattern_sheet(s)]:
        df = sheet_df(wb_vor_lin, sname, "Fleet")
        col = find_col(df, ["Fleet No", "FleetNo", "Fleet#"])
        if col: vor_set |= set(df[col].map(normalize).values) - {""}

    for sname in wb_vor_lin.sheetnames:
        if "disposal" not in sname.lower(): continue
        df = sheet_df(wb_vor_lin, sname, "Fleet")
        col = find_col(df, ["Fleet No", "FleetNo", "Fleet#"])
        if col: dispose_set |= set(df[col].map(normalize).values) - {""}

    # Trucker dict: fleet# → depot ถ้า PTD == today
    trucker_depot = {}
    frames = []
    for sname in wb_trucker.sheetnames:
        df = sheet_df(wb_trucker, sname, "Trailer")
        if not df.empty: frames.append(df)

    if frames:
        tdf = pd.concat(frames, ignore_index=True)
        tc        = find_col(tdf, ["Trailer"])
        conf_dep  = find_col(tdf, ["Confirmed Depart DC Time", "Confirmed Depart DC"])
        ptd       = find_col(tdf, ["Depart DC PTD"])
        depc      = find_col(tdf, ["Depart Depot", "Depot"])
        if tc and depc:
            tdf["_fn"] = tdf[tc].map(normalize)
            # เช็ค Confirmed Depart DC Time ก่อน ถ้าว่างหรือไม่ตรงให้ใช้ Depart DC PTD
            def matches_today(row, target_date):
                # เช็ค Confirmed Depart DC Time ก่อน
                if conf_dep and conf_dep in row.index:
                    val = row[conf_dep]
                    if val is not None and str(val).strip() not in ("", "NaT", "None", "nan"):
                        # มีค่า → ใช้ค่านี้ตัดสินเลย ไม่ fallback
                        d = pd.to_datetime(val, dayfirst=True, errors="coerce")
                        return pd.notna(d) and d.date() == target_date
                # ว่างเท่านั้น → fallback ไป Depart DC PTD
                if ptd and ptd in row.index:
                    val = row[ptd]
                    if val is not None and str(val).strip() not in ("", "NaT", "None", "nan"):
                        d = pd.to_datetime(val, dayfirst=True, errors="coerce")
                        if pd.notna(d) and d.date() == target_date:
                            return True
                return False
            tdf["_is_match"] = tdf.apply(lambda r: matches_today(r, today), axis=1)
            today_rows = tdf[tdf["_is_match"]]
            for fn_val, grp in today_rows.groupby("_fn"):
                if not fn_val: continue
                last_row = grp.iloc[-1]
                depot = str(last_row[depc]).strip()
                trucker_depot[fn_val] = depot if depot else "Unknown Depot"

    return vor_set, dispose_set, trucker_depot


def get_status_for_trailer(fn, vor_set, dispose_set, trucker_depot):
    fn = normalize(fn)
    if not fn: return "N/A"
    if fn in vor_set:       return "VOR"
    if fn in dispose_set:   return "Dispose"
    if fn in trucker_depot: return trucker_depot[fn]
    return "N/A"


@st.cache_data(show_spinner=False)
def process_all_files(lotus_bytes, vor_all_bytes, vor_lin_bytes, trucker_bytes, run_date, gps_bytes, trucker2_bytes):
    import io as _io
    wb_lotus   = openpyxl.load_workbook(_io.BytesIO(lotus_bytes),   data_only=True)
    wb_vor_all = openpyxl.load_workbook(_io.BytesIO(vor_all_bytes), data_only=True)
    wb_vor_lin = openpyxl.load_workbook(_io.BytesIO(vor_lin_bytes), data_only=True)
    wb_trucker = openpyxl.load_workbook(_io.BytesIO(trucker_bytes), data_only=True)

    # Read Lotus
    lotus_ws  = wb_lotus[wb_lotus.sheetnames[0]]
    hdr_row   = find_header_row(lotus_ws, "Fleet")
    data      = list(lotus_ws.values)
    cols      = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(data[hdr_row])]
    lotus_df  = pd.DataFrame(data[hdr_row + 1:], columns=cols)
    fleet_col = find_col(lotus_df, ["Fleet#", "Fleet No", "Fleet Number", "Trailer", "Registration"])
    if fleet_col is None:
        return None, None, "❌ ไม่พบ column Fleet# ใน Trailer Lotus"

    # Build lookup sets 1 ครั้ง แล้วค้นหา O(1)
    vor_set, dispose_set, trucker_depot = build_lookups(wb_vor_all, wb_vor_lin, wb_trucker, run_date)
    statuses = [
        get_status_for_trailer(val, vor_set, dispose_set, trucker_depot)
        for val in lotus_df[fleet_col]
    ]
    lotus_df["🚦 Status"] = statuses
    lotus_df = lotus_df.reset_index(drop=True)

    # Trucker ไฟล์ที่ 2 (optional) — ตรวจทับทุกสถานะที่ไม่ใช่ชื่อ Depot
    DEPOT_STATUSES = {"VOR", "Dispose", "On Road", "N/A", ""}
    if trucker2_bytes:
        try:
            wb_trucker2 = openpyxl.load_workbook(_io.BytesIO(trucker2_bytes), data_only=True)
            t2_frames = []
            for sname in wb_trucker2.sheetnames:
                df2 = sheet_df(wb_trucker2, sname, "Trailer")
                if df2.empty: continue
                df2["_sheet"] = sname
                t2_frames.append(df2)
            if t2_frames:
                t2_df = pd.concat(t2_frames, ignore_index=True)
                t2_trailer = find_col(t2_df, ["Trailer"])
                t2_depot   = find_col(t2_df, ["Depart Depot", "Depot"])
                t2_conf_dep = find_col(t2_df, ["Confirmed Depart DC Time", "Confirmed Depart DC"])
                t2_ptd2     = find_col(t2_df, ["Depart DC PTD"])
                if t2_trailer and t2_depot:
                    t2_df["_fn"] = t2_df[t2_trailer].map(normalize)
                    for i, row in lotus_df.iterrows():
                        current_status = str(lotus_df.at[i, "🚦 Status"]).strip()
                        is_depot = (current_status not in DEPOT_STATUSES
                                    and not current_status.startswith("Parking")
                                    and not current_status.startswith("Locked")
                                    and current_status != "On Road")
                        if is_depot:
                            continue
                        fn_val = normalize(row[fleet_col])
                        matched2 = t2_df[t2_df["_fn"] == fn_val].iloc[::-1]
                        for _, row2 in matched2.iterrows():
                            is_match = False
                            # เช็ค Confirmed Depart DC Time ก่อน
                            if t2_conf_dep and t2_conf_dep in row2.index:
                                val = row2[t2_conf_dep]
                                if val is not None and str(val).strip() not in ("", "NaT", "None", "nan"):
                                    # มีค่า → ใช้ค่านี้ตัดสินเลย ไม่ fallback
                                    d = pd.to_datetime(val, errors='coerce')
                                    is_match = pd.notna(d) and d.date() == run_date
                                    if not is_match:
                                        break  # มีค่าแต่ไม่ตรง → ข้ามแถวนี้
                            else:
                                # ว่างเท่านั้น → fallback ไป Depart DC PTD
                                if t2_ptd2 and t2_ptd2 in row2.index:
                                    val = row2[t2_ptd2]
                                    if val is not None and str(val).strip() not in ("", "NaT", "None", "nan"):
                                        d = pd.to_datetime(val, errors='coerce')
                                        if pd.notna(d) and d.date() == run_date:
                                            is_match = True
                            if is_match:
                                depot2 = str(row2[t2_depot]).strip()
                                lotus_df.at[i, "🚦 Status"] = depot2 if depot2 else "Unknown Depot"
                                break
        except Exception as e:
            pass

    # GPS
    gps_fleet_set = set()
    gps_lookup = {}
    if gps_bytes:
        try:
            wb_gps = openpyxl.load_workbook(_io.BytesIO(gps_bytes), data_only=True)
            gps_df = sheet_df(wb_gps, wb_gps.sheetnames[0], "ชื่อรถ")
            col_car     = find_col(gps_df, ["ชื่อรถ"])
            col_status  = find_col(gps_df, ["สถานะ"])
            col_station = find_col(gps_df, ["ชื่อสถานี"])
            if col_car:
                gps_fleet_set = set(gps_df[col_car].map(normalize).values) - {""}
            if col_car and col_station:
                for _, row in gps_df.iterrows():
                    fn_gps  = normalize(row[col_car])
                    station = str(row[col_station]).strip()
                    if not fn_gps: continue
                    abbr = next((code for kw, code in STATION_MAP.items() if kw.upper() in station.upper()), None)
                    if abbr:
                        gps_lookup[fn_gps] = f"Parking{abbr}"
                    else:
                        gps_lookup[fn_gps] = "On Road"
        except Exception:
            pass

    # Apply Trailer Rental (เฉพาะ N/A)
    type_col     = find_col(lotus_df, ["Type"])
    owner_dc_col = find_col(lotus_df, ["OwnerDC", "Owner DC"])
    if type_col and owner_dc_col:
        for i, row in lotus_df.iterrows():
            if lotus_df.at[i, "🚦 Status"] != "N/A": continue
            fn_val   = normalize(row[fleet_col])
            type_val = str(row[type_col]).strip()
            owner_dc = str(row[owner_dc_col]).strip()
            if type_val == "Trailer Rental":
                lotus_df.at[i, "🚦 Status"] = f"Locked {owner_dc}" if fn_val in gps_fleet_set else ""

    # Apply GPS Parking (เฉพาะ N/A)
    for i, fn in enumerate(lotus_df[fleet_col]):
        if lotus_df.at[i, "🚦 Status"] == "N/A":
            fnu = normalize(fn)
            if fnu in gps_lookup:
                lotus_df.at[i, "🚦 Status"] = gps_lookup[fnu]

    return lotus_df, fleet_col, None


def style_status(val):
    s = str(val).strip().upper()
    if s == "VOR":     return "background-color:#FEE2E2; color:#991B1B; font-weight:700; border-radius:6px;"
    if s == "DISPOSE": return "background-color:#F3E8FF; color:#6B21A8; font-weight:700; border-radius:6px;"
    if s == "ON ROAD": return "background-color:#DCFCE7; color:#166534; font-weight:700; border-radius:6px;"
    if s == "N/A":     return "background-color:#F1F5F9; color:#94A3B8; border-radius:6px;"
    if s == "":        return "background-color:#F1F5F9; color:#94A3B8; border-radius:6px;"
    if s.startswith("PARKING"): return "background-color:#FEF9C3; color:#854D0E; font-weight:700; border-radius:6px;"
    if s.startswith("LOCKED"):  return "background-color:#FDE68A; color:#92400E; font-weight:700; border-radius:6px;"
    return "background-color:#DBEAFE; color:#1E40AF; font-weight:700; border-radius:6px;"

# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

@st.dialog("✅ ยืนยันวันที่ประมวลผล")
def confirm_dialog(run_date):
    st.markdown(f"""
    <div style="text-align:center; padding: 8px 0 16px;">
        <div style="font-size:3rem; margin-bottom:8px;">📅</div>
        <div style="font-size:1rem; color:#475569; margin-bottom:4px;">วันที่ที่เลือกสำหรับการประมวลผล</div>
        <div style="font-size:1.8rem; font-weight:800; color:#1B4FD8; letter-spacing:-0.02em;">
            {run_date.strftime('%d %B %Y')}
        </div>
    </div>
    """, unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ ยืนยัน", use_container_width=True, type="primary"):
            st.session_state.confirmed = True
            st.rerun()
    with col2:
        if st.button("❌ ยกเลิก", use_container_width=True):
            st.session_state.confirmed = False
            st.rerun()


st.markdown("""
<div class="dash-header">
  <div class="dash-header-eyebrow">Fleet Intelligence Platform</div>
  <h1>Trailer Status Dashboard</h1>
  <p>ระบบตรวจสอบสถานะตู้รถแบบ Real-time | Lotus / Linfox Fleet Management</p>
  <div class="dash-header-badge"><span class="dot"></span> Live Dashboard</div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## 📂 อัพโหลดไฟล์ Excel")
    st.markdown("---")
    st.markdown('<div class="upload-box"><h4>1️⃣ Trailer Lotus *</h4><div class="upload-desc">รายการทะเบียนตู้รถทั้งหมด</div>', unsafe_allow_html=True)
    f_lotus = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="lotus")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:40px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-box"><h4>2️⃣ VOR Allnow *</h4><div class="upload-desc">ตู้ชำรุด / รอขาย (Allnow)</div>', unsafe_allow_html=True)
    f_vor_all = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="vor_all")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:40px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-box"><h4>3️⃣ VOR Linfox *</h4><div class="upload-desc">ตู้ชำรุด / รอขาย (Linfox)</div>', unsafe_allow_html=True)
    f_vor_lin = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="vor_lin")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:40px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-box"><h4>4️⃣ Trucker *</h4><div class="upload-desc">ข้อมูล Trip วันปัจจุบัน</div>', unsafe_allow_html=True)
    f_trucker = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="trucker")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:40px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-box" style="background:#F1F5F9; border-color:#CBD5E1; border-style:dashed;"><h4 style="color:#64748B !important;">5️⃣ Trucker วันถัดไป</h4><div class="upload-desc">ข้อมูล Trip วันถัดไป (optional)</div>', unsafe_allow_html=True)
    f_trucker2 = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="trucker2")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:40px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-box"><h4>6️⃣ GPS *</h4><div class="upload-desc">ข้อมูลตำแหน่งรถ</div>', unsafe_allow_html=True)
    f_gps = st.file_uploader("", label_visibility="collapsed", type=["xlsx","xlsm"], key="gps")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")
    run_date = st.date_input("📅 วันที่ประมวลผล", value=date.today())
    
    # คำเตือนวันที่
    today_actual = date.today()
    if run_date == today_actual:
        st.markdown(f'<div class="custom-success">✅ วันที่ถูกต้อง: {run_date.strftime("%d %b %Y")}</div>', unsafe_allow_html=True)
    else:
        diff = (today_actual - run_date).days
        if diff > 0:
            st.markdown(f'<div class="custom-warning">⚠️ วันที่ย้อนหลัง {diff} วัน<br>{run_date.strftime("%d %b %Y")}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="custom-warning">⚠️ วันที่ล่วงหน้า {abs(diff)} วัน<br>{run_date.strftime("%d %b %Y")}</div>', unsafe_allow_html=True)

    if "confirmed" not in st.session_state:
        st.session_state.confirmed = False

    open_btn = st.button("▶️  ประมวลผล", use_container_width=True, type="primary")
    if open_btn:
        confirm_dialog(run_date)

    run_btn = st.session_state.confirmed

if not run_btn:
    st.info("📌 กรุณาอัพโหลดไฟล์ Excel ทั้ง 4 ไฟล์ในแถบด้านซ้าย แล้วกด **ประมวลผล**")
    st.stop()

# reset confirmed หลังผ่าน dialog แล้ว
st.session_state.confirmed = False

missing = []
if not f_lotus:   missing.append("Trailer Lotus")
if not f_vor_all: missing.append("VOR Allnow")
if not f_vor_lin: missing.append("VOR Linfox")
if not f_trucker: missing.append("Trucker")
if not f_gps:     missing.append("GPS")
if missing:
    st.error(f"❌ ยังไม่ได้อัพโหลดไฟล์: **{', '.join(missing)}**")
    st.stop()

with st.spinner("กำลังโหลดและประมวลผล..."):
    try:
        lotus_bytes    = f_lotus.read()
        vor_all_bytes  = f_vor_all.read()
        vor_lin_bytes  = f_vor_lin.read()
        trucker_bytes  = f_trucker.read()
        trucker2_bytes = f_trucker2.read() if f_trucker2 else b""
        gps_bytes      = f_gps.read() if f_gps else b""
    except Exception as e:
        st.error(f"❌ อ่านไฟล์ไม่สำเร็จ: {e}")
        st.stop()

lotus_df, fleet_col, err = process_all_files(
    lotus_bytes, vor_all_bytes, vor_lin_bytes, trucker_bytes, run_date, gps_bytes, trucker2_bytes
)
if err:
    st.error(err)
    st.stop()

statuses = lotus_df["🚦 Status"].tolist()
total       = len(lotus_df)
vor_cnt     = sum(1 for s in statuses if s == "VOR")
dispose_cnt = sum(1 for s in statuses if s == "Dispose")
depot_cnt   = sum(1 for s in statuses if s not in ("VOR","Dispose","On Road","N/A","") and not str(s).startswith("Parking") and not str(s).startswith("Locked"))
parking_cnt = sum(1 for s in statuses if str(s).startswith("Parking"))
onroad_cnt  = sum(1 for s in statuses if s == "On Road")
locked_cnt  = sum(1 for s in statuses if str(s).startswith("Locked"))
na_cnt      = sum(1 for s in statuses if s == "N/A")
blank_cnt   = sum(1 for s in statuses if s == "")

cols_m = st.columns(8)
def metric_card(col, val, label, color):
    col.markdown(f"""<div class="metric-card">
      <div class="accent-bar" style="background:{color};"></div>
      <div class="val" style="color:{color}">{val}</div>
      <div class="lbl">{label}</div>
    </div>""", unsafe_allow_html=True)

metric_card(cols_m[0], total,       "🚛 ทั้งหมด",     "#1e3a5f")
metric_card(cols_m[1], vor_cnt,     "⚠️ VOR",          "#dc2626")
metric_card(cols_m[2], dispose_cnt, "🗑️ Dispose",      "#6b21a8")
metric_card(cols_m[3], depot_cnt,   "🏭 ออกเดินทาง",  "#2563eb")
metric_card(cols_m[4], parking_cnt, "🅿️ Parking",      "#d97706")
metric_card(cols_m[5], onroad_cnt,  "🟢 On Road",      "#16a34a")
metric_card(cols_m[6], locked_cnt,  "🔒 Locked",       "#92400e")
metric_card(cols_m[7], blank_cnt,   "— ว่าง",          "#94A3B8")

st.markdown("<br>", unsafe_allow_html=True)
st.markdown("<h3 style='color:#0F172A;font-weight:700;font-size:1.1rem;margin-bottom:4px;'>📋 รายการตู้รถทั้งหมด</h3>", unsafe_allow_html=True)

display_df = lotus_df.copy()
st.markdown(f"<p class='count-label'>แสดง <b>{len(display_df):,}</b> รายการ</p>", unsafe_allow_html=True)

# แทนช่องว่างด้วย "—" ในตาราง
display_df["🚦 Status"] = display_df["🚦 Status"].replace("", "—")

styled = (
    display_df.style
    .map(style_status, subset=["🚦 Status"])
    .set_properties(**{"font-size": "13px"})
    .set_table_styles([
        {"selector": "thead th", "props": [("background-color","#1e3a5f"),("color","white"),("font-size","13px"),("font-weight","700"),("padding","10px 14px")]},
        {"selector": "tbody tr:nth-of-type(even)", "props": [("background-color","#f8fafc")]},
        {"selector": "tbody td", "props": [("padding","8px 14px")]},
    ])
)
st.dataframe(styled, use_container_width=True, height=520)

with st.expander("📖 คำอธิบายสถานะ"):
    st.markdown("""
| สถานะ | ความหมาย |
|---|---|
| ⚠️ **VOR** | ตู้รถชำรุด |
| 🗑️ **Dispose** | รอขาย |
| 🏭 **ชื่อ Depot** | ออกเดินทางวันนี้ ยังไม่มีเวลากลับ |
| 🅿️ **ParkingXXDC** | จอดอยู่ที่ DC ตามข้อมูล GPS |
| 🚛 **On Road** | อยู่ระหว่างเส้นทาง |
| 🔒 **Locked** | ตู้เช่าที่พบใน GPS |
| ⚪ **N/A** | ไม่พบข้อมูลในระบบ |
""")

st.markdown("---")
output = io.BytesIO()
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    display_df.to_excel(writer, index=False, sheet_name="Trailer Status")
output.seek(0)
st.download_button(
    label="⬇️ ดาวน์โหลดผลลัพธ์ (.xlsx)",
    data=output,
    file_name=f"trailer_status_{run_date.strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
st.markdown(f"<p style='text-align:center;color:#334155;font-size:.72rem;letter-spacing:.05em;'>TRAILER STATUS DASHBOARD · ประมวลผล {run_date.strftime('%d %B %Y')} · v2.0</p>", unsafe_allow_html=True)
